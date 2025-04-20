# interacting with an excel workbook in python
# import modules
import os, sys
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, Font, Border, Side

# a function to load excel into a pandas dataframe
# dataframes are often used to manipulate data 
def open_excel_workbook(workbook_path:str, target_worksheet_name:str):
    '''
    a function for reading an excel document

    PARAMETERS
    ----------
    workbook_path : str - path to excel workbook
    target_worksheet_name : str - excel worksheet name

    RETURN 
    ------
    dataframe : pd.Dataframe - pandas dataframe

    '''

    # check excel file exists
    if not os.path.isfile(workbook_path):
        error_message = f"{workbook_path} not found"
        raise FileNotFoundError(error_message)

    # load excel workbook as an object
    workbook = load_workbook(filename=workbook_path)

    # check worksheets
    worksheets = workbook.sheetnames

    if len(worksheets) == 0:
        error_message = f"{workbook_path} is empty"
        raise ValueError(error_message)

    # output results to terminal
    print(f"workbook {workbook_path} found with {worksheet} worksheets")

    # check target worksheet exists
    if target_worksheet_name not in worksheets:
        error_message = f"{target_worksheet_name} not found in worksheet list {worksheets} for document {workbook_path}"
        raise ValueError(error_message)

    # load target worksheet into a dataframe
    target_worksheet = pd.read_excel(workbook_path, sheet_name=target_worksheet_name)

    # return target worksheet if it not empty
    return None if target_worksheet.empty else target_worksheet

class kf_workbooks():

    def __init__(self, path:str):
        '''
        a class to access excel workbooks

        PARAMETERS
        -----------
        path : str - a path to a workbook

        '''
        
        # check the path is a string
        # we raise an error if the path is not a string and highlight what datatype it is
        if not isinstance(path, str):
            error_message = f"path {type(path)} is not a string"
            raise ValueError(error_message)
        
        # check the path exists
        # we raise an error if the path does not exist
        if not os.path.exists(path):
            error_message = f"path {path} not found"
            raise ValueError(error_message)
        
        # define the path
        self.path = path

    def __str__(self):
        return f"kf_workbook - accessing: {self.path}"
    
    def openpyxl(self, worksheet_name:str=None):
        '''
        load workbook into an openpyxl object

        PARAMETERS
        -----------
        worksheet_name : str - a workbook tab name
        '''

        workbook = load_workbook(self.path)

        if worksheet_name != None:
            worksheets = workbook.sheetnames
            
            # check the worksheet name exists in the workbook
            if worksheet_name not in worksheets:
                error_message = f"worksheet {worksheet_name} not in {worksheets} {self.path}"
                raise ValueError(error_message)
            
            # return a worksheet object rather than a workbook object
            worksheet = workbook[worksheet_name]
            return worksheet
        
        # return a workbook if no worksheet name is provided
        return workbook
    
    def dataframe(self, worksheet_name:str):
        '''
        load workbook into a pandas dataframe

        PARAMETERS
        ----------
        worksheet : str - a worksheet name

        '''

        # load excel workbook as an object
        workbook = load_workbook(self.path)

        # check worksheets
        worksheets = workbook.sheetnames

        if len(worksheets) == 0:
            error_message = f"{workbook_path} is empty"
            raise ValueError(error_message)

        # output results to terminal
        print(f"workbook {workbook_path} found with {worksheet} worksheets")

        # check target worksheet exists
        if worksheet_name not in worksheets:
            error_message = f"{worksheet_name} not found in worksheet list {worksheets} for document {workbook_path}"
            raise ValueError(error_message)

        # load target worksheet into a dataframe
        target_worksheet = pd.read_excel(workbook_path, sheet_name=worksheet_name)

        # return target worksheet if it not empty
        return None if target_worksheet.empty else target_worksheet
    

if __name__ == "__main__":

    # a test to turn the worksheet 'shopping" from the workbook 'data.xlsx' into a dataframe
    workbook_path = "data.xlsx"
    worksheet = "shopping"

    # open the workbook into a dataframe using a function
    dataframe = open_excel_workbook(workbook_path, worksheet)
    print(f"dataframe\n{dataframe}")

    # an example of filtering the dataframe
    filtered_dataframe = dataframe[dataframe['item'] == 'orange']
    print(f"filtered dataframe\n{filtered_dataframe}")

    # open the workbook into a dataframe using a class
    result = kf_workbooks(workbook_path).openpyxl()
    print(f"class returned\n{result}")
